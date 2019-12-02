import openpyxl
import xlrd
import os
from .exceptions import FailToLoadFile


class ExcelProcessor:
    """
    The class holds API function for processing and handling an excel file.

    todo: find a better name or try to merge it into the the parser class.
    """

    def __init__(self, file_path, logger):
        """
        Init the function.

        :param file_path: The path of the file to process.
        :param logger: The logger object.
        """
        self.logger = logger
        self._file_path = file_path
        self.sheet_names = []

        # Load file.
        if not self._load_excel(file_path=self._file_path):
            raise FailToLoadFile(f"Failed to load excel file - {self._file_path}")

    def _load_excel(self, file_path):
        """
        Load excel file.

        :param file_path: full path.
        :return:
        """
        if not os.path.exists(file_path):
            self.logger.error(f"File not exists {file_path}")
            return False

        filename, extension = os.path.splitext(file_path)

        if extension.lower() == '.xlsx':

            try:
                # Load in the workbook file.
                self._workbook = openpyxl.load_workbook(filename=file_path)
            except Exception as e:
                self.logger.error(f"Failed to load xlsx file - {str(e)}, {file_path}")
                return False

        elif extension.lower() == '.xls':

            try:
                self._workbook = xlrd.open_workbook(filename=file_path)
            except Exception as e:
                self.logger.error(f"Failed to load xls file -  {str(e)}, {file_path}")
                return False

        if not self._workbook:
            self.logger(f"Failed to load excel file {file_path}")
            return False

        self.sheet_names = self._workbook.sheetnames

        return True

    def get_cell(self, sheet_name, row, column):
        """
        Get cell value.

        :param sheet_name: The sheet name.
        :param row: The row.
        :param column: column.

        :return:
        """
        try:
            if sheet_name not in self.sheet_names:
                self.logger.warn("sheet name not exists in excel")
                return None

            row = self._workbook[sheet_name].cell(row=row, column=column).value

            if not row:
                return None

            try:
                return str(row)
            except ValueError:
                return row

        except Exception as ex:
            raise Exception(f"Failed to read cell {ex}")

    def get_entire_row(self, sheet_name, row, min_column=1, max_column=None):
        """
        Get row between min column number to max column number. if max column is None, get all cells until cell data is
        None.

        :param sheet_name:
        :param row:
        :param min_column:
        :param max_column:

        :return: row data :type: list
        """

        if sheet_name not in self.sheet_names:
            self.logger.warn(f"sheet name, {sheet_name}, does not exists")
            return None

        row_data = []
        column = min_column

        # Get cell data.
        cell_data = self.get_cell(sheet_name=sheet_name, column=column, row=row)

        while self._should_iterate_columns(max_column, column, cell_data):
            row_data.append(cell_data)
            column += 1
            cell_data = self.get_cell(sheet_name=sheet_name, column=column, row=row)

        return row_data

    def _should_iterate_columns(self, max_column, column, cell_data):
        """
        If max column than use is_not_max_column lambda to check if is the max column. If max column is None, use
        data_exists lambda to check if cell data exists.

        :param max_column: Maximum columns we have.
        :param column: The current column indicator.
        :param cell_data: The data of the cell.

        :return: Rather we need to iterate over the current column or not.
        """
        if max_column:
            return not(max_column == column)

        if cell_data:
            return True

        return False

