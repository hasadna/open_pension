from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import re

p = re.compile('[0-9]{4}')
anchor_cell_value = 'מזומנים ושווי מזומנים'
months = {
    1: 'ינואר',
    2: 'פברואר',
    3: 'מרץ',
    4: 'אפריל',
    5: 'מאי',
    6: 'יוני',
    7: 'יולי',
    8: 'אוגוסט',
    9: 'ספטמבר',
    10: 'נובמבר',
    11: 'אוקטובר',
    12: 'דצמבר',
}


def process_file(file: Workbook):
    """
    Handling a file for process.

    :param file: The file to process.
    """
    sheets = {}
    for worksheet in file.worksheets:
        sheets[worksheet.title] = process_data(worksheet)

    return sheets


def is_year(cell_value) -> bool:
    """
    Getting the year.
    """

    if p.match(str(cell_value)):
        return True

    return False


def collect_table(worksheet: Worksheet, year: int, fund_name, row: int, column: int) -> object:
    """
    Start to run on the fields and look for the anchor cell which from there
    the data lay out.
    """

    data = {}

    for iterated_row in range(row, row + 21):
        row_label: str = worksheet.cell(row=iterated_row, column=column).value

        row_data = {}
        for iterated_column in range(column + 1, column + 1 + 24):

            # Getting the matching month of the current cell and resetting
            # the cell value.
            month_index = int(iterated_column / 2)
            month_name = months[month_index]

            if month_name not in row_data:
                row_data[month_name] = {
                    'התרומה לתשואה': '',
                    'שיעור מסך הנכסים': '',
                    'year': year,
                    'fund': fund_name,
                }

            cell_kwargs = {'row': iterated_row, 'column': iterated_column}
            cell_value = worksheet.cell(**cell_kwargs).value

            if cell_value:
                cell_value = format(cell_value * 100, '.2f')

            # Values in cell with even index number have a label and odd
            # indexed-cell have a different label.
            if iterated_column % 2 == 0:
                key = 'התרומה לתשואה'
            else:
                key = 'שיעור מסך הנכסים'
            row_data[month_name][key] = cell_value

        data[row_label.strip()] = row_data

    return data


def remove_none_fund_label(optional_labels):
    needle = None
    for optional_label in optional_labels:

        if is_year(optional_label):
            # ￿Skip the year.
            continue

        if 'ההשקעה לתשואה הכוללת' in optional_label:
            # This one if not the fund label any way.
            continue

        needle = optional_label

    return needle


def process_data(worksheet: Worksheet) -> (int, object):
    # Resetting the variables: the year and variables which will help us to
    # collect the fund name.
    year = 0
    fund_name = None
    optional_fund_name = []

    for row in range(worksheet.max_row):
        row_position = row + 1
        for column in range(worksheet.max_column):
            # Bump the row and the column by 1 since the API method which
            # returns the value does not allow 0 as parameters.
            column_position = column + 1
            cell_value = worksheet.cell(row_position, column_position).value

            if cell_value is None:
                continue

            if fund_name is None:
                optional_fund_name.append(cell_value)

            # First, let's get the year.
            if is_year(cell_value):
                year = cell_value

                # We found the year which means some of the items in the
                # optional_fund_name could be the one. Start by filter the year:
                fund_name = remove_none_fund_label(optional_fund_name)

            # After we got the year, let's look for the anchor cell.
            if cell_value == anchor_cell_value:
                collect_table_kwargs = {
                    'worksheet': worksheet,
                    'year': year,
                    'fund_name': fund_name,
                    'row': row_position,
                    'column': column_position,
                }
                return collect_table(**collect_table_kwargs)
