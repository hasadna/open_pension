import openpyxl
import xlrd
import os
import datetime


class ExcelLoader:
    def __init__(self, file_path, logger):
        self._logger = logger
        self._file_path = file_path
        self.sheet_names = []

        # load file
        if not self._load_excel(file_path=self._file_path):
            raise Exception("Failed to load excel file")

    def _load_excel(self, file_path):
        """
        Load excel file
        :param file_path: full path
        :return:
        """
        if not os.path.exists(file_path):
            self._logger.error("File not exists {0}".format(file_path))
            return False
        if os.path.splitext(file_path)[1].lower() == '.xlsx':
            try:
                # Load in the workbook file
                self._workbook = openpyxl.load_workbook(filename=file_path)
            except Exception as ex:
                self._logger.error("Failed to load xlsx file -  {0}, {1} ".format(ex, file_path))
                return False
        elif os.path.splitext(file_path)[1].lower() == '.xls':
            try:
                self._workbook = xlrd.open_workbook(filename=file_path)
                # TODO: adjust code to work with xlrd
            except Exception as ex:
                self._logger.error("Failed to load xls file -  {0}, {1} ".format(ex, file_path))
                return False

        if not self._workbook:
            self._logger("Failed to load excel file {0}".format(file_path))
            return False

        self.sheet_names = self._workbook.sheetnames

        return True

    def get_cell(self, sheet_name, row, column):
        """
        Get cell value
        :param sheet_name:
        :param row:
        :param column:
        :return:
        """
        try:
            if sheet_name not in self.sheet_names:
                self._logger.warn("sheet name not exists in excel")
                return None

            row = self._workbook[sheet_name].cell(row=row, column=column).value

            # if isinstance(row, datetime.datetime):
            #     return str(row)

            if not row:
                return None

            try:
                return str(row)
            except ValueError:
                return row

        except Exception as ex:
            raise Exception("Failed to read cell {0}".format(ex))

    def get_entire_row(self, sheet_name, row, min_column=1, max_column=None):
        """
        Get row between min column number to max column number
        if max column is None, get all cells until cell data is None
        :param sheet_name:
        :param row:
        :param min_column:
        :param max_column:
        :return: row data :type: list
        """

        if sheet_name not in self.sheet_names:
            self._logger.warn("sheet name not exists in excel")
            return None

        cell_data = None
        row_data = []
        column = min_column

        # lambdas function
        data_exists = lambda: True if cell_data else False
        is_not_max_column = lambda: not(max_column == column)

        # If max column than use is_not_max_column lambda to check if is the max column
        # If max column is None, use data_exists lambda to check if cell data exists
        if max_column:
            check = is_not_max_column
        else:
            check = data_exists

        # Get cell data
        cell_data = self.get_cell(sheet_name=sheet_name, column=column, row=row)

        while check():
            row_data.append(cell_data)
            column += 1
            cell_data = self.get_cell(sheet_name=sheet_name, column=column, row=row)

        return row_data

#
# if __name__ == '__main__':
#     # For debug only
#     class FakeLogger:
#         def error(self, msg):
#             print("error {0}".format(msg))
#
#         def info(self, msg):
#             print("info {0}".format(msg))
#
#         def warn(self, msg):
#             print("warring {0}".format(msg))
#
#     logger = FakeLogger()
#
#     excel = ExcelLoader(file_path="/tmp/s1.xlsx", logger=logger)
#     print(excel.get_entire_row(sheet_name="מניות", row=1, min_column=2))